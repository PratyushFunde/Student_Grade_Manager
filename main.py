from openpyxl import Workbook, load_workbook
import os

file_path = 'Students.xlsx'

def load_or_create_workbook(file_path):
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()
    return wb

def add_data(ws):
    num=count_non_empty_cells(ws)
    l = []
    names=[]
    name=input("Enter name of student : ")
    names.append(name)
    
    
    for i in range(num-1):
        n = int(input("Enter marks : "))
        l.append(n)
    names=names+l
    ws.append(names)
    print("Data added successfully.")

def view_data(ws):
    for row in ws.iter_rows(values_only=True):
        print(row)

def create_new_worksheet(wb):
    sheet_name = input("Enter the name of the new worksheet: ")
    
    if sheet_name in wb.sheetnames:
        print(f"A worksheet with the name '{sheet_name}' already exists.")
    else:
        ws = wb.create_sheet(title=sheet_name)
        print(f"Worksheet '{sheet_name}' created successfully.")
        headings=[]
        print("----------You can enter headings only once----------")
        n=int(input("Enter number of headings : "))

        for i in range(n):
            h=input(f"Enter heading {i+1} : ")
            headings.append(h)
        ws=wb[sheet_name]
        ws.append(headings)
    # return ws


def delete_sheet(wb):
   
    sheet_name=input("Enter worksheet you want to delete : ")
   
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
        wb.save("Students.xlsx")
    else:
        print(f"\n!!! The {sheet_name} worksheet do not exits !!!")
       
def count_non_empty_cells(ws):
    non_empty_cells = 0
   
    first_row = ws[1]  # Access the first row
    first_row_values = [cell.value for cell in first_row]
    return len(first_row_values)

def main():
    while True:
        print("\nOptions:")
        print("1. Add data")
        print("2. View data")
        print("3. Create new worksheet")
        print("4. To delete worksheet")
        print("0. Exit")
        choice = input("Enter your choice: ")

        if choice == '0':
            break
        elif choice == '1':

            wb = load_or_create_workbook(file_path)
            print("Available worksheets are : ",wb.sheetnames)
            sheet_name = input("Enter the name of the worksheet to add data: ")
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                add_data(ws)
                wb.save(file_path)
            else:
                print(f"No worksheet named '{sheet_name}' exists.")


        elif choice == '2':

            wb = load_or_create_workbook(file_path)
            print(f"Available sheets are : {wb.sheetnames}")
            sheet_name = input("Enter the name of the worksheet to view data: ")
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                view_data(ws)
            else:
                print(f" !!!-----No worksheet named '{sheet_name}'-----!!!exists.")

        elif choice == '3':

            wb = load_or_create_workbook(file_path)
            print("Available worksheets are ",wb.sheetnames)
            ws = create_new_worksheet(wb)
            wb.save(file_path)

        elif choice == '4':
              
              wb=load_or_create_workbook(file_path)
            #   print(wb)
              print("----------These are available sheets----------")
              print(wb.sheetnames)
              print(len(wb.sheetnames))
              if(len(wb.sheetnames)==1):
                  print('''
                        You can't delete the sheet as there is only one sheet,
                        Create one more sheet and then delete this sheet
                        ''')
              else:
                delete_sheet(wb)  
        else:
            print("Invalid choice. Please try again.")

if __name__ == "__main__":
    print("Main")
    main()
    